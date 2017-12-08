from openpyxl import Workbook
from openpyxl import load_workbook
import re
targets=['investment', 'merge', 'm&a', 'announcement']
def load_dataset(filename):
        wb=load_workbook(filename)
        data=wb.active
        return data
def searchListWord(string,ls):
        if string==None:
		return 	 False
        string=string.lower()
        for l  in ls:
                if string.find(l)!=-1:
                        return True

        return False
def dict2list(dic:dict):

    keys = dic.keys()
    vals = dic.values()
    lst = [(key, val) for key, val in zip(keys, vals)]
    return lst
def xlsx2txt(texts):
        #res={}
        respre={}
        inform={}
        for text in texts:
                txt=text.value
                if searchListWord(txt,targets):

                        txt=txt.lower()
                        txt=re.sub(r'\d'," ",txt)
                        txt=re.findall(r'\w+(?:[\&]\w+)*',txt)
                        txt=(' '.join(txt))
                        res={}

                        maxtar=0
                        for target in targets:
                                countertarget=txt.count(target)
                                res[target]=countertarget
                                maxtar+=countertarget
                        respre[str(text.row)]=maxtar
                        inform[text]=res
                        
                
        a=sorted(dict2list(respre),key=lambda x:x[1],reverse=True)
        nums=[]
        for i in range(len(a)):
                nums.append( a[i][0])
        return nums
def writexlsx(filename,nums):
        wb=Workbook()
        sheet=wb.active
        sheet.title="result"

        i=1
        for num in nums:
                j=1
                datas=load_dataset(filename)[num]
                for data in datas:
                        #print(data.value)
                        sheet.cell(row=i,column=j,value=str( data.value))
                        j=j+1
                i=i+1
        wb.save('result.xlsx')
if __name__=='__main__':
        filename='scraped-text (1).xlsx'

        texts=load_dataset(filename)['H']
        nums=xlsx2txt(texts)
        writexlsx(filename,nums)

